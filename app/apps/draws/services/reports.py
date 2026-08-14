from datetime import timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile
from zoneinfo import ZoneInfo

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.files import File
from django.db.models import Count
from django.db.models.functions import TruncDate
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from draws.models import Draw, DrawReport, Winner
from promo.models import PromoCode, PromoCodeAttempt


HEADER_FILL = PatternFill('solid', fgColor='FFD329')
HEADER_FONT = Font(bold=True, color='111111')


def generate_draw_report(report_id):
    report = DrawReport.objects.get(pk=report_id)
    date_from, date_to = _resolve_period(report)
    campaign_timezone = ZoneInfo(settings.TIME_ZONE)

    users_by_day = _counts_by_day(
        get_user_model().objects.filter(
            date_joined__date__range=(date_from, date_to),
        ),
        'date_joined',
        campaign_timezone,
    )
    codes_by_day = _counts_by_day(
        PromoCode.objects.filter(
            registered_at__date__range=(date_from, date_to),
        ),
        'registered_at',
        campaign_timezone,
    )
    attempts_by_day = _attempt_counts(
        date_from,
        date_to,
        campaign_timezone,
    )
    draws = list(
        Draw.objects.filter(draw_date__range=(date_from, date_to))
        .prefetch_related('winners')
        .order_by('draw_date')
    )
    draws_by_day = {draw.draw_date: draw for draw in draws}
    participant_counts = {
        draw.pk: _participant_count(draw) for draw in draws
    }

    workbook = Workbook()
    daily_sheet = workbook.active
    daily_sheet.title = 'По дням'
    draws_sheet = workbook.create_sheet('Розыгрыши')
    winners_sheet = workbook.create_sheet('Победители')

    _write_daily_sheet(
        daily_sheet,
        date_from,
        date_to,
        users_by_day,
        codes_by_day,
        attempts_by_day,
        draws_by_day,
        participant_counts,
    )
    _write_draws_sheet(draws_sheet, draws, participant_counts)
    _write_winners_sheet(winners_sheet, date_from, date_to)

    with NamedTemporaryFile(suffix='.xlsx') as temporary_file:
        workbook.save(temporary_file.name)
        filename = f'draw-report-{date_from}-{date_to}.xlsx'
        with Path(temporary_file.name).open('rb') as report_file:
            report.report_file.save(
                filename,
                File(report_file),
                save=False,
            )
        report.date_from = date_from
        report.date_to = date_to
        report.save(update_fields=('date_from', 'date_to', 'report_file'))

    return filename


def _resolve_period(report):
    date_from = report.date_from
    date_to = report.date_to or timezone.localdate()

    if date_from is None:
        candidates = []
        first_user = (
            get_user_model().objects.order_by('date_joined')
            .values_list('date_joined', flat=True)
            .first()
        )
        first_code = (
            PromoCode.objects.filter(registered_at__isnull=False)
            .order_by('registered_at')
            .values_list('registered_at', flat=True)
            .first()
        )
        first_attempt = (
            PromoCodeAttempt.objects.order_by('created_at')
            .values_list('created_at', flat=True)
            .first()
        )
        first_draw = (
            Draw.objects.order_by('draw_date')
            .values_list('draw_date', flat=True)
            .first()
        )
        for value in (first_user, first_code, first_attempt):
            if value is not None:
                candidates.append(timezone.localtime(value).date())
        if first_draw is not None:
            candidates.append(first_draw)
        date_from = min(candidates, default=date_to)

    return date_from, date_to


def _counts_by_day(queryset, field_name, campaign_timezone):
    rows = (
        queryset.annotate(
            report_date=TruncDate(field_name, tzinfo=campaign_timezone),
        )
        .values('report_date')
        .annotate(total=Count('pk'))
    )
    return {row['report_date']: row['total'] for row in rows}


def _attempt_counts(date_from, date_to, campaign_timezone):
    rows = (
        PromoCodeAttempt.objects.filter(
            created_at__date__range=(date_from, date_to),
        )
        .annotate(
            report_date=TruncDate('created_at', tzinfo=campaign_timezone),
        )
        .values('report_date', 'result')
        .annotate(total=Count('pk'))
    )
    result = {}
    for row in rows:
        counts = result.setdefault(
            row['report_date'],
            {'successful': 0, 'unsuccessful': 0},
        )
        if row['result'] == PromoCodeAttempt.Result.SUCCESS:
            counts['successful'] += row['total']
        else:
            counts['unsuccessful'] += row['total']
    return result


def _participant_count(draw):
    if not draw.period_started_at or not draw.period_ended_at:
        return 0

    candidates = PromoCode.objects.filter(
        registered_at__gte=draw.period_started_at,
        registered_at__lt=draw.period_ended_at,
        registered_by__isnull=False,
    )
    if draw.started_at:
        previous_winner_ids = Winner.objects.filter(
            won_at__lt=draw.started_at,
        ).values_list('user_id', flat=True)
        candidates = candidates.exclude(registered_by_id__in=previous_winner_ids)
    return candidates.values('registered_by_id').distinct().count()


def _write_daily_sheet(
    sheet,
    date_from,
    date_to,
    users_by_day,
    codes_by_day,
    attempts_by_day,
    draws_by_day,
    participant_counts,
):
    _append_header(
        sheet,
        (
            'Дата',
            'Новые пользователи',
            'Зарегистрированные промокоды',
            'Успешные попытки',
            'Неуспешные попытки',
            'Тип запуска',
            'Участники',
            'Победители',
        ),
    )
    current_date = date_from
    while current_date <= date_to:
        draw = draws_by_day.get(current_date)
        attempt_counts = attempts_by_day.get(
            current_date,
            {'successful': 0, 'unsuccessful': 0},
        )
        sheet.append(
            (
                current_date,
                users_by_day.get(current_date, 0),
                codes_by_day.get(current_date, 0),
                attempt_counts['successful'],
                attempt_counts['unsuccessful'],
                draw.get_trigger_display() if draw else '',
                participant_counts.get(draw.pk, 0) if draw else 0,
                draw.winners.count() if draw else 0,
            )
        )
        current_date += timedelta(days=1)
    sheet.auto_filter.ref = sheet.dimensions
    _set_widths(sheet, (14, 22, 31, 21, 24, 20, 15, 15))


def _write_draws_sheet(sheet, draws, participant_counts):
    _append_header(
        sheet,
        (
            'Дата',
            'Статус',
            'Тип запуска',
            'Начало периода',
            'Конец периода',
            'Запущен',
            'Завершён',
            'Участники',
            'Победители',
        ),
    )
    for draw in draws:
        sheet.append(
            (
                draw.draw_date,
                draw.get_status_display(),
                draw.get_trigger_display(),
                _format_datetime(draw.period_started_at),
                _format_datetime(draw.period_ended_at),
                _format_datetime(draw.started_at),
                _format_datetime(draw.completed_at),
                participant_counts.get(draw.pk, 0),
                draw.winners.count(),
            )
        )
    sheet.auto_filter.ref = sheet.dimensions
    _set_widths(sheet, (14, 18, 20, 24, 24, 24, 24, 15, 15))


def _write_winners_sheet(sheet, date_from, date_to):
    _append_header(
        sheet,
        (
            'Дата розыгрыша',
            'Тип запуска',
            'Email',
            'Фамилия',
            'Имя',
            'Отчество',
            'Телефон',
            'Приз',
            'Промокод',
            'Время победы',
            'Письмо отправлено',
        ),
    )
    winners = (
        Winner.objects.filter(draw__draw_date__range=(date_from, date_to))
        .select_related('draw', 'user', 'user__profile', 'promo_code')
        .order_by('draw__draw_date', 'prize')
    )
    for winner in winners:
        profile = getattr(winner.user, 'profile', None)
        sheet.append(
            (
                winner.draw.draw_date,
                winner.draw.get_trigger_display(),
                winner.user.email,
                profile.last_name if profile else '',
                profile.first_name if profile else '',
                profile.middle_name if profile else '',
                profile.phone if profile else '',
                winner.get_prize_display(),
                winner.promo_code.code,
                _format_datetime(winner.won_at),
                _format_datetime(winner.notified_at),
            )
        )
    sheet.auto_filter.ref = sheet.dimensions
    _set_widths(sheet, (20, 20, 32, 22, 22, 22, 22, 20, 15, 24, 24))


def _append_header(sheet, values):
    sheet.append(values)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    sheet.freeze_panes = 'A2'


def _set_widths(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def _format_datetime(value):
    if value is None:
        return ''
    return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S')
