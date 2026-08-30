from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from threading import Barrier
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import call, patch
from zoneinfo import ZoneInfo

from django.core import mail
from django.core.cache import cache
from django.core.management import call_command
from django.contrib.auth.models import Permission
from django.db import connections, IntegrityError, transaction
from django.test import TestCase, TransactionTestCase
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import Profile, User
from draws.models import Draw, DrawReport, Winner
from draws.services.draw import (
    InvalidDrawPeriod,
    get_pending_draw_dates,
    run_draw,
)
from draws.services.notifications import send_winner_email
from draws.services.public_results import PUBLIC_DRAWS_CACHE_KEY
from draws.tasks import (
    cleanup_expired_report_files_task,
    generate_draw_report_task,
    run_daily_draw_task,
)
from promo.models import PromoCode, PromoCodeAttempt
from promo.services.registration import register_promo_code


class DrawsAdminTests(TestCase):
    def setUp(self):
        admin_user = User.objects.create_superuser(
            email='admin@example.com',
            password='test-password',
        )
        self.client.force_login(admin_user)

    def test_draw_admin_pages_render(self):
        urls = (
            reverse('admin:draws_draw_changelist'),
            reverse('admin:draws_winner_changelist'),
        )

        for url in urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)

    def test_dashboard_does_not_contain_draw_shortcuts(self):
        response = self.client.get(reverse('admin:index'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Быстрые действия')

    def test_report_dialog_uses_native_date_inputs(self):
        response = self.client.get(
            reverse('admin:draws_draw_generate_report'),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'type="date"', count=2)

    @patch('draws.admin.timezone.now')
    @patch('draws.admin.run_manual_draw_task.delay')
    def test_starts_manual_draw_from_admin(self, delay, now):
        cutoff = datetime(
            2026, 8, 13, 9, 30, tzinfo=ZoneInfo('Europe/Moscow')
        )
        now.return_value = cutoff
        delay.return_value = SimpleNamespace(id='manual-task-id')

        response = self.client.post(
            reverse('admin:draws_draw_run_manual_draw'),
            {'_form_submitted': True, 'confirmation': 'on'},
        )

        self.assertRedirects(
            response,
            reverse('admin:draws_draw_changelist'),
        )
        delay.assert_called_once_with(
            '2026-08-13',
            cutoff.isoformat(),
        )

    @patch('draws.admin.timezone.now')
    @patch('draws.admin.run_manual_draw_task.delay')
    def test_manual_draw_dialog_redirects_with_htmx(self, delay, now):
        cutoff = datetime(
            2026, 8, 13, 9, 30, tzinfo=ZoneInfo('Europe/Moscow')
        )
        now.return_value = cutoff
        delay.return_value = SimpleNamespace(id='manual-task-id')

        response = self.client.post(
            reverse('admin:draws_draw_run_manual_draw'),
            {'_form_submitted': True, 'confirmation': 'on'},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 204)
        self.assertEqual(
            response.headers['HX-Redirect'],
            reverse('admin:draws_draw_changelist'),
        )

    def test_view_permission_cannot_start_manual_draw(self):
        viewer = User.objects.create_user(
            email='viewer@example.com',
            password='test-password',
            is_staff=True,
        )
        viewer.user_permissions.add(
            Permission.objects.get(
                content_type__app_label='draws',
                codename='view_draw',
            )
        )
        self.client.force_login(viewer)

        response = self.client.post(
            reverse('admin:draws_draw_run_manual_draw'),
            {'_form_submitted': True, 'confirmation': 'on'},
        )

        self.assertEqual(response.status_code, 403)

    @patch('draws.admin.generate_draw_report_task.delay')
    def test_queues_xlsx_report_from_admin(self, delay):
        delay.return_value = SimpleNamespace(id='report-task-id')

        response = self.client.post(
            reverse('admin:draws_draw_generate_report'),
            {
                '_form_submitted': True,
                'date_from': '2026-08-01',
                'date_to': '2026-08-13',
            },
        )

        report = DrawReport.objects.get()
        self.assertRedirects(
            response,
            reverse('admin:draws_drawreport_changelist'),
        )
        self.assertEqual(report.date_from.isoformat(), '2026-08-01')
        self.assertEqual(report.date_to.isoformat(), '2026-08-13')
        self.assertEqual(report.celery_task_id, 'report-task-id')
        delay.assert_called_once_with(report.pk)

    @patch('draws.admin.generate_draw_report_task.delay')
    def test_report_dialog_redirects_htmx_to_report_log(self, delay):
        delay.return_value = SimpleNamespace(id='report-task-id')

        response = self.client.post(
            reverse('admin:draws_draw_generate_report'),
            {
                '_form_submitted': True,
                'date_from': '',
                'date_to': '',
            },
            headers={'HX-Request': 'true'},
        )

        self.assertEqual(response.status_code, 204)
        self.assertEqual(
            response.headers['HX-Redirect'],
            reverse('admin:draws_drawreport_changelist'),
        )
        self.assertTrue(DrawReport.objects.exists())

    @patch('draws.admin.generate_draw_report_task.delay')
    def test_rejects_report_with_reversed_period(self, delay):
        response = self.client.post(
            reverse('admin:draws_draw_generate_report'),
            {
                '_form_submitted': True,
                'date_from': '2026-08-13',
                'date_to': '2026-08-01',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'Дата начала не может быть позже даты окончания.',
        )
        self.assertFalse(DrawReport.objects.exists())
        delay.assert_not_called()


class DrawReportTests(TestCase):
    report_date = datetime(2026, 8, 12).date()
    moscow = ZoneInfo('Europe/Moscow')

    def setUp(self):
        self.media_directory = TemporaryDirectory()
        settings_override = override_settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        settings_override.enable()
        self.addCleanup(settings_override.disable)
        self.addCleanup(self.media_directory.cleanup)
        self.admin_user = User.objects.create_superuser(
            email='reporter@example.com',
            password='test-password',
        )

    def create_report_data(self):
        participant = User.objects.create_user(
            email='winner@example.com',
            password='StrongPassword_123!',
        )
        Profile.objects.create(
            user=participant,
            first_name='Иван',
            last_name='Иванов',
            middle_name='Иванович',
            phone='+79990000000',
        )
        registration_time = datetime(
            2026, 8, 12, 12, 0, tzinfo=self.moscow,
        )
        User.objects.filter(pk=participant.pk).update(
            date_joined=registration_time,
        )
        promo_code = PromoCode.objects.create(
            code='REPORT01',
            registered_by=participant,
            registered_at=registration_time,
        )
        attempt = PromoCodeAttempt.objects.create(
            user=participant,
            raw_code=promo_code.code,
            normalized_code=promo_code.code,
            result=PromoCodeAttempt.Result.SUCCESS,
            reason=PromoCodeAttempt.Reason.SUCCESS,
        )
        PromoCodeAttempt.objects.filter(pk=attempt.pk).update(
            created_at=registration_time,
        )
        draw = Draw.objects.create(
            draw_date=self.report_date,
            status=Draw.Status.COMPLETED,
            trigger=Draw.Trigger.MANUAL,
            period_started_at=datetime(
                2026, 8, 12, 0, 0, tzinfo=self.moscow,
            ),
            period_ended_at=datetime(
                2026, 8, 12, 18, 0, tzinfo=self.moscow,
            ),
            started_at=datetime(
                2026, 8, 12, 18, 0, tzinfo=self.moscow,
            ),
            completed_at=datetime(
                2026, 8, 12, 18, 1, tzinfo=self.moscow,
            ),
        )
        Winner.objects.create(
            draw=draw,
            user=participant,
            promo_code=promo_code,
            prize=Winner.Prize.AIRPODS,
        )
        return participant

    def test_generates_daily_draw_and_winner_sheets(self):
        participant = self.create_report_data()
        report = DrawReport.objects.create(
            date_from=self.report_date,
            date_to=self.report_date,
            created_by=self.admin_user,
        )

        generate_draw_report_task(report.pk)

        report.refresh_from_db()
        self.assertEqual(report.status, DrawReport.Status.COMPLETED)
        self.assertTrue(report.report_file)
        with report.report_file.open('rb') as report_file:
            workbook = load_workbook(report_file, read_only=True, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                ['По дням', 'Розыгрыши', 'Победители'],
            )
            daily_row = list(
                workbook['По дням'].iter_rows(min_row=2, values_only=True)
            )[0]
            winner_row = list(
                workbook['Победители'].iter_rows(
                    min_row=2,
                    values_only=True,
                )
            )[0]
            workbook.close()

        self.assertEqual(daily_row[1:5], (1, 1, 1, 0))
        self.assertEqual(daily_row[6:8], (1, 1))
        self.assertEqual(winner_row[2], participant.email)
        self.assertEqual(
            winner_row[3:7],
            ('Иванов', 'Иван', 'Иванович', '+79990000000'),
        )
        self.assertEqual(winner_row[8], 'REPORT01')

    @override_settings(GENERATED_FILE_RETENTION_DAYS=7)
    def test_cleanup_deletes_report_file_but_keeps_history(self):
        self.create_report_data()
        report = DrawReport.objects.create(
            date_from=self.report_date,
            date_to=self.report_date,
            created_by=self.admin_user,
        )
        generate_draw_report_task(report.pk)
        DrawReport.objects.filter(pk=report.pk).update(
            created_at=timezone.now() - timedelta(days=8),
        )

        deleted_count = cleanup_expired_report_files_task()

        report.refresh_from_db()
        self.assertEqual(deleted_count, 1)
        self.assertFalse(report.report_file)
        self.assertTrue(DrawReport.objects.filter(pk=report.pk).exists())


class PublicDrawApiTests(TestCase):
    def setUp(self):
        cache.clear()

    def test_lists_completed_draws_without_private_user_data(self):
        user = User.objects.create_user(
            email='winner@example.com',
            password='StrongPassword_123!',
        )
        Profile.objects.create(
            user=user,
            first_name='Михаил',
            last_name='Иванов',
            phone='+7 (999) 123-45-67',
            no_middle_name=True,
        )
        promo_code = PromoCode.objects.create(
            code='PUBLIC01',
            registered_by=user,
            registered_at=timezone.now(),
        )
        draw = Draw.objects.create(
            draw_date=datetime(2026, 8, 12).date(),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        Winner.objects.create(
            draw=draw,
            user=user,
            promo_code=promo_code,
            prize=Winner.Prize.AIRPODS,
        )
        Draw.objects.create(
            draw_date=datetime(2026, 8, 11).date(),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        Draw.objects.create(
            draw_date=datetime(2026, 8, 13).date(),
            status=Draw.Status.PENDING,
        )

        response = self.client.get(reverse('public-draw-list'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 2)
        self.assertEqual(response.json()[0]['draw_date'], '2026-08-12')
        self.assertEqual(
            response.json()[0]['winners'],
            [
                {
                    'name': 'Михаил И.',
                    'prize_code': Winner.Prize.AIRPODS,
                    'prize_name': 'Наушники AirPods',
                }
            ],
        )
        self.assertEqual(response.json()[1]['winners'], [])
        payload = str(response.json())
        self.assertNotIn(user.email, payload)
        self.assertNotIn(Profile.objects.get(user=user).phone, payload)
        self.assertNotIn(promo_code.code, payload)

    def test_invalidates_cached_payload_after_winner_commit(self):
        user = User.objects.create_user(
            email='winner@example.com',
            password='StrongPassword_123!',
        )
        profile = Profile.objects.create(
            user=user,
            first_name='Михаил',
            last_name='Иванов',
        )
        draw = Draw.objects.create(
            draw_date=datetime(2026, 8, 12).date(),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        promo_code = PromoCode.objects.create(
            code='PUBLIC01',
            registered_by=user,
            registered_at=timezone.now(),
        )

        first_response = self.client.get(reverse('public-draw-list'))
        self.assertEqual(first_response.json()[0]['winners'], [])
        self.assertIsNotNone(cache.get(PUBLIC_DRAWS_CACHE_KEY))

        with self.captureOnCommitCallbacks(execute=True):
            Winner.objects.create(
                draw=draw,
                user=user,
                promo_code=promo_code,
                prize=Winner.Prize.AIRPODS,
            )
        self.assertIsNone(cache.get(PUBLIC_DRAWS_CACHE_KEY))

        second_response = self.client.get(reverse('public-draw-list'))
        self.assertEqual(
            second_response.json()[0]['winners'][0]['name'],
            'Михаил И.',
        )

        profile.first_name = 'Анна'
        with self.captureOnCommitCallbacks(execute=True):
            profile.save(update_fields=('first_name',))
        self.assertIsNone(cache.get(PUBLIC_DRAWS_CACHE_KEY))

        third_response = self.client.get(reverse('public-draw-list'))
        self.assertEqual(
            third_response.json()[0]['winners'][0]['name'],
            'Анна И.',
        )


class DrawServiceTests(TransactionTestCase):
    draw_date = datetime(2026, 8, 12).date()
    moscow = ZoneInfo('Europe/Moscow')

    def setUp(self):
        self.notification_patcher = patch(
            'draws.services.draw.schedule_winner_emails'
        )
        self.schedule_winner_emails = self.notification_patcher.start()
        self.addCleanup(self.notification_patcher.stop)

    def create_participant(self, index, registered_at=None):
        user = User.objects.create_user(
            email=f'user{index}@example.com',
            password='StrongPassword_123!',
            is_email_verified=True,
        )
        Profile.objects.create(user=user)
        promo_code = PromoCode.objects.create(
            code=f'CD{index:06d}',
            registered_by=user,
            registered_at=(
                registered_at
                or datetime(2026, 8, 12, 12, 0, tzinfo=self.moscow)
            ),
        )
        return user, promo_code

    @patch('draws.services.draw.secrets.randbelow', return_value=0)
    def test_selects_two_winners_from_draw_date(self, random_offset):
        participants = [self.create_participant(index) for index in range(3)]

        draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.AUTOMATIC,
        )

        draw.refresh_from_db()
        winners = list(draw.winners.order_by('prize'))
        self.assertEqual(draw.status, Draw.Status.COMPLETED)
        self.assertEqual(len(winners), 2)
        self.assertEqual(
            {winner.prize for winner in winners},
            {Winner.Prize.OZON_3000, Winner.Prize.AIRPODS},
        )
        self.assertEqual(len({winner.user_id for winner in winners}), 2)
        self.assertTrue(
            {winner.user_id for winner in winners}.issubset(
                {user.pk for user, _ in participants}
            )
        )
        self.assertEqual(random_offset.call_count, 2)

    @patch('draws.services.draw.secrets.randbelow', return_value=0)
    def test_excludes_past_winner_and_same_user_codes(self, random_offset):
        past_winner, past_code = self.create_participant(
            1,
            datetime(2026, 8, 11, 12, 0, tzinfo=self.moscow),
        )
        past_draw = Draw.objects.create(
            draw_date=self.draw_date - timedelta(days=1),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        Winner.objects.create(
            draw=past_draw,
            user=past_winner,
            promo_code=past_code,
            prize=Winner.Prize.OZON_3000,
        )
        PromoCode.objects.create(
            code='PAST0001',
            registered_by=past_winner,
            registered_at=datetime(
                2026, 8, 12, 10, 0, tzinfo=self.moscow
            ),
        )
        first_user, _ = self.create_participant(2)
        PromoCode.objects.create(
            code='SAME0002',
            registered_by=first_user,
            registered_at=datetime(
                2026, 8, 12, 14, 0, tzinfo=self.moscow
            ),
        )
        second_user, _ = self.create_participant(3)

        draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.AUTOMATIC,
        )

        winner_user_ids = set(
            draw.winners.values_list('user_id', flat=True)
        )
        self.assertEqual(winner_user_ids, {first_user.pk, second_user.pk})
        self.assertNotIn(past_winner.pk, winner_user_ids)

    def test_manual_cutoff_excludes_later_registration(self):
        cutoff = datetime(2026, 8, 12, 15, 0, tzinfo=self.moscow)
        eligible_user, _ = self.create_participant(
            1,
            cutoff - timedelta(minutes=1),
        )
        late_user, _ = self.create_participant(
            2,
            cutoff + timedelta(minutes=1),
        )

        draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.MANUAL,
            cutoff=cutoff,
        )

        self.assertEqual(draw.winners.count(), 1)
        self.assertEqual(draw.winners.get().user, eligible_user)
        self.assertNotEqual(draw.winners.get().user, late_user)
        self.assertEqual(
            draw.period_started_at,
            datetime(2026, 8, 12, 0, 0, tzinfo=self.moscow),
        )
        self.assertEqual(draw.period_ended_at, cutoff)

    @patch('draws.services.draw.secrets.randbelow', return_value=0)
    def test_next_draw_uses_codes_registered_after_manual_draw(
        self,
        random_offset,
    ):
        manual_cutoff = datetime(
            2026, 8, 12, 12, 0, tzinfo=self.moscow
        )
        self.create_participant(1, manual_cutoff - timedelta(minutes=1))
        manual_draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.MANUAL,
            cutoff=manual_cutoff,
        )
        late_user, late_code = self.create_participant(
            2,
            manual_cutoff + timedelta(minutes=1),
        )

        next_draw = run_draw(
            draw_date=self.draw_date + timedelta(days=1),
            trigger=Draw.Trigger.AUTOMATIC,
        )

        self.assertEqual(manual_draw.period_ended_at, manual_cutoff)
        self.assertEqual(next_draw.period_started_at, manual_cutoff)
        self.assertEqual(next_draw.winners.count(), 1)
        winner = next_draw.winners.get()
        self.assertEqual(winner.user, late_user)
        self.assertEqual(winner.promo_code, late_code)

    def test_repeat_run_returns_existing_result(self):
        self.create_participant(1)
        self.create_participant(2)
        cutoff = datetime(2026, 8, 12, 23, 0, tzinfo=self.moscow)
        first_draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.MANUAL,
            cutoff=cutoff,
        )
        winner_ids = list(
            first_draw.winners.order_by('pk').values_list('pk', flat=True)
        )

        repeated_draw = run_draw(
            draw_date=self.draw_date,
            trigger=Draw.Trigger.AUTOMATIC,
        )

        self.assertEqual(repeated_draw.pk, first_draw.pk)
        self.assertEqual(
            list(
                repeated_draw.winners.order_by('pk').values_list(
                    'pk', flat=True
                )
            ),
            winner_ids,
        )
        repeated_draw.refresh_from_db()
        self.assertEqual(repeated_draw.trigger, Draw.Trigger.MANUAL)
        self.assertEqual(self.schedule_winner_emails.call_count, 2)

    def test_finds_every_uncompleted_date_since_first_registration(self):
        self.create_participant(
            1,
            datetime(2026, 8, 10, 12, 0, tzinfo=self.moscow),
        )
        Draw.objects.create(
            draw_date=date(2026, 8, 10),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )

        pending_dates = get_pending_draw_dates(
            through_date=date(2026, 8, 13),
        )

        self.assertEqual(
            pending_dates,
            [date(2026, 8, 11), date(2026, 8, 12), date(2026, 8, 13)],
        )

    def test_rejects_naive_or_invalid_cutoff(self):
        with self.assertRaises(InvalidDrawPeriod):
            run_draw(
                draw_date=self.draw_date,
                trigger=Draw.Trigger.MANUAL,
                cutoff=datetime(2026, 8, 12, 12, 0),
            )

        with self.assertRaises(InvalidDrawPeriod):
            run_draw(
                draw_date=self.draw_date,
                trigger=Draw.Trigger.MANUAL,
                cutoff=datetime(
                    2026, 8, 12, 0, 0, tzinfo=self.moscow
                ),
            )

    def test_concurrent_runs_create_one_draw_and_one_set_of_winners(self):
        self.create_participant(1)
        self.create_participant(2)
        self.create_participant(3)
        barrier = Barrier(2)

        def start_draw():
            connections.close_all()
            try:
                barrier.wait(timeout=5)
                return run_draw(
                    draw_date=self.draw_date,
                    trigger=Draw.Trigger.AUTOMATIC,
                ).pk
            finally:
                connections.close_all()

        with ThreadPoolExecutor(max_workers=2) as executor:
            draw_ids = [
                future.result(timeout=10)
                for future in (
                    executor.submit(start_draw),
                    executor.submit(start_draw),
                )
            ]

        self.assertEqual(len(set(draw_ids)), 1)
        self.assertEqual(Draw.objects.count(), 1)
        self.assertEqual(Winner.objects.count(), 2)

    def test_draw_and_registration_do_not_deadlock(self):
        user, _ = self.create_participant(1)
        profile = user.profile
        profile.first_name = 'Иван'
        profile.last_name = 'Иванов'
        profile.no_middle_name = True
        profile.phone = '+7 (999) 123-45-67'
        profile.save()
        code_to_register = PromoCode.objects.create(code='REGISTER')
        barrier = Barrier(2)

        def start_draw():
            connections.close_all()
            try:
                barrier.wait(timeout=5)
                return run_draw(
                    draw_date=self.draw_date,
                    trigger=Draw.Trigger.AUTOMATIC,
                ).pk
            finally:
                connections.close_all()

        def register_code():
            connections.close_all()
            try:
                barrier.wait(timeout=5)
                return register_promo_code(
                    user=user,
                    raw_code=code_to_register.code,
                ).pk
            finally:
                connections.close_all()

        with ThreadPoolExecutor(max_workers=2) as executor:
            draw_future = executor.submit(start_draw)
            registration_future = executor.submit(register_code)
            draw_future.result(timeout=10)
            registration_future.result(timeout=10)

        code_to_register.refresh_from_db()
        self.assertEqual(code_to_register.registered_by_id, user.pk)


class DrawConstraintTests(TransactionTestCase):
    def test_rejects_invalid_draw_status_and_trigger(self):
        with self.assertRaises(IntegrityError), transaction.atomic():
            Draw.objects.create(
                draw_date=datetime(2026, 8, 12).date(),
                status='SUCCES',
            )

        with self.assertRaises(IntegrityError), transaction.atomic():
            Draw.objects.create(
                draw_date=datetime(2026, 8, 12).date(),
                trigger='CELRY',
            )

    def test_rejects_invalid_prize(self):
        user = User.objects.create_user(
            email='winner@example.com',
            password='StrongPassword_123!',
        )
        promo_code = PromoCode.objects.create(
            code='WINN0001',
            registered_by=user,
            registered_at=timezone.now(),
        )
        draw = Draw.objects.create(
            draw_date=timezone.localdate(),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            Winner.objects.create(
                draw=draw,
                user=user,
                promo_code=promo_code,
                prize='Ozon3000',
            )


class DailyDrawTaskTests(TestCase):
    @patch('draws.tasks.get_pending_draw_dates')
    @patch('draws.tasks.run_draw')
    @patch('draws.tasks.timezone.now')
    def test_catches_up_every_missed_date(
        self,
        now,
        draw_service,
        pending_draw_dates,
    ):
        now.return_value = datetime(
            2026,
            8,
            11,
            21,
            0,
            tzinfo=ZoneInfo('UTC'),
        )
        pending_draw_dates.return_value = [
            date(2026, 8, 10),
            date(2026, 8, 11),
        ]
        draw_service.side_effect = [
            SimpleNamespace(
                pk=41,
                draw_date=date(2026, 8, 10),
                winners=SimpleNamespace(count=lambda: 2),
            ),
            SimpleNamespace(
                pk=42,
                draw_date=date(2026, 8, 11),
                winners=SimpleNamespace(count=lambda: 1),
            ),
        ]

        result = run_daily_draw_task()

        pending_draw_dates.assert_called_once_with(
            through_date=date(2026, 8, 11),
        )
        self.assertEqual(
            draw_service.call_args_list,
            [
                call(
                    draw_date=date(2026, 8, 10),
                    trigger=Draw.Trigger.AUTOMATIC,
                ),
                call(
                    draw_date=date(2026, 8, 11),
                    trigger=Draw.Trigger.AUTOMATIC,
                ),
            ],
        )
        self.assertEqual(
            result,
            {
                'draw_count': 2,
                'draws': [
                    {
                        'draw_id': 41,
                        'draw_date': '2026-08-10',
                        'winner_count': 2,
                    },
                    {
                        'draw_id': 42,
                        'draw_date': '2026-08-11',
                        'winner_count': 1,
                    },
                ],
            },
        )

    @patch('draws.management.commands.run_missed_draws.run_daily_draw_task')
    def test_management_command_runs_catch_up_until_given_date(self, task):
        task.return_value = {'draw_count': 0, 'draws': []}

        call_command('run_missed_draws', '--through', '2026-08-13')

        task.assert_called_once_with(date(2026, 8, 13))

    @override_settings(TIME_ZONE='Europe/Moscow')
    def test_beat_registers_daily_draw_task(self):
        from django.conf import settings

        entry = settings.CELERY_BEAT_SCHEDULE[
            'run-daily-draw-at-moscow-midnight'
        ]

        self.assertEqual(entry['task'], 'draws.tasks.run_daily_draw_task')
        self.assertEqual(entry['options']['expires'], 3600)


class WinnerNotificationTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            email='winner@example.com',
            password='StrongPassword_123!',
        )
        Profile.objects.create(
            user=self.user,
            promo_code_email_notifications=False,
        )
        self.promo_code = PromoCode.objects.create(
            code='PRIZE001',
            registered_by=self.user,
            registered_at=timezone.now(),
        )
        self.draw = Draw.objects.create(
            draw_date=timezone.localdate(),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
        )
        self.winner = Winner.objects.create(
            draw=self.draw,
            user=self.user,
            promo_code=self.promo_code,
            prize=Winner.Prize.AIRPODS,
        )

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'
    )
    def test_sends_mandatory_winner_email_and_marks_notification(self):
        sent = send_winner_email(self.winner.pk)

        self.assertTrue(sent)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, [self.user.email])
        self.assertIn('AirPods', mail.outbox[0].body)
        self.assertIn(self.promo_code.code, mail.outbox[0].body)
        self.winner.refresh_from_db()
        self.assertIsNotNone(self.winner.notified_at)

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'
    )
    def test_does_not_send_winner_email_twice(self):
        self.assertTrue(send_winner_email(self.winner.pk))
        self.assertFalse(send_winner_email(self.winner.pk))

        self.assertEqual(len(mail.outbox), 1)
