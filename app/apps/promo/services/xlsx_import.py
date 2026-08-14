import re
from pathlib import Path
from tempfile import NamedTemporaryFile

from django.core.files import File
from django.db import connection, transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from promo.models import PromoCode, PromoCodeImport
from promo.services.generation import promo_code_generation_lock


BATCH_SIZE = 5_000
CODE_PATTERN = re.compile(r'^[A-Z0-9]{8}$')


def import_promo_codes(import_id):
    import_job = PromoCodeImport.objects.get(pk=import_id)
    seen_codes = set()
    pending_rows = []
    processed_count = 0
    imported_count = 0
    skipped_count = 0

    error_report = _ErrorReport()

    with import_job.source_file.open('rb') as source:
        workbook = load_workbook(source, read_only=True, data_only=True)
        worksheet = workbook.active

        try:
            with promo_code_generation_lock():
                for row_number, (cell,) in enumerate(
                    worksheet.iter_rows(min_col=1, max_col=1, values_only=True),
                    start=1,
                ):
                    if cell is None or str(cell).strip() == '':
                        continue

                    processed_count += 1
                    raw_value = str(cell).strip()
                    code = raw_value.upper()

                    if not CODE_PATTERN.fullmatch(code):
                        skipped_count += 1
                        error_report.append(
                            (row_number, raw_value, 'Неверный формат')
                        )
                    elif code in seen_codes:
                        skipped_count += 1
                        error_report.append(
                            (row_number, raw_value, 'Дубликат в файле')
                        )
                    else:
                        seen_codes.add(code)
                        pending_rows.append((row_number, raw_value, code))

                    if len(pending_rows) >= BATCH_SIZE:
                        imported, skipped = _insert_batch(
                            pending_rows,
                            error_report,
                        )
                        imported_count += imported
                        skipped_count += skipped
                        pending_rows.clear()

                    if processed_count % BATCH_SIZE == 0:
                        _update_progress(
                            import_id,
                            processed_count,
                            imported_count,
                            skipped_count,
                        )

                if pending_rows:
                    imported, skipped = _insert_batch(
                        pending_rows,
                        error_report,
                    )
                    imported_count += imported
                    skipped_count += skipped
        finally:
            workbook.close()

    if skipped_count:
        _save_error_workbook(import_job, error_report.workbook)

    _update_progress(
        import_id,
        processed_count,
        imported_count,
        skipped_count,
    )
    return {
        'processed_count': processed_count,
        'imported_count': imported_count,
        'skipped_count': skipped_count,
    }


def _insert_batch(rows, error_report):
    codes = [code for _, _, code in rows]
    existing_codes = set(
        PromoCode.objects.filter(code__in=codes).values_list('code', flat=True)
    )
    new_rows = [row for row in rows if row[2] not in existing_codes]

    for row_number, raw_value, code in rows:
        if code in existing_codes:
            error_report.append(
                (row_number, raw_value, 'Код уже существует в базе')
            )

    inserted_codes = _insert_codes([row[2] for row in new_rows])
    for row_number, raw_value, code in new_rows:
        if code not in inserted_codes:
            error_report.append(
                (row_number, raw_value, 'Код уже существует в базе')
            )

    skipped_count = len(rows) - len(inserted_codes)
    return len(inserted_codes), skipped_count


def _insert_codes(codes):
    if not codes:
        return set()

    table_name = connection.ops.quote_name(PromoCode._meta.db_table)
    placeholders = ', '.join(['(%s, %s)'] * len(codes))
    created_at = timezone.now()
    parameters = []
    for code in codes:
        parameters.extend((code, created_at))

    query = (
        f'INSERT INTO {table_name} ("code", "created_at") '
        f'VALUES {placeholders} '
        'ON CONFLICT ("code") DO NOTHING RETURNING "code"'
    )

    with transaction.atomic(), connection.cursor() as cursor:
        cursor.execute(query, parameters)
        return {row[0] for row in cursor.fetchall()}


def _update_progress(import_id, processed, imported, skipped):
    PromoCodeImport.objects.filter(pk=import_id).update(
        processed_count=processed,
        imported_count=imported,
        skipped_count=skipped,
    )


def _save_error_workbook(import_job, workbook):
    with NamedTemporaryFile(suffix='.xlsx') as temporary_file:
        workbook.save(temporary_file.name)
        filename = f'import-{import_job.pk}-errors.xlsx'
        with Path(temporary_file.name).open('rb') as error_file:
            import_job.error_file.save(
                filename,
                File(error_file),
                save=False,
            )
        import_job.save(update_fields=('error_file',))


class _ErrorReport:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def append(self, row):
        if self.workbook is None:
            self.workbook = Workbook(write_only=True)
            self.sheet = self.workbook.create_sheet('Пропущенные строки')
            self.sheet.column_dimensions['A'].width = 12
            self.sheet.column_dimensions['B'].width = 24
            self.sheet.column_dimensions['C'].width = 34
            self.sheet.append(('Строка', 'Значение', 'Причина'))
        self.sheet.append(row)
