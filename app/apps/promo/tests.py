from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from io import BytesIO
from threading import Barrier
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch

from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connections, IntegrityError
from django.test import Client, override_settings, TestCase, TransactionTestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from accounts.models import Profile, User
from accounts.services.authentication import create_token_pair
from draws.models import Draw, Winner
from promo.models import (
    PromoCode,
    PromoCodeAttempt,
    PromoCodeGeneration,
    PromoCodeImport,
)
from promo.services.notifications import send_registration_email
from promo.tasks import (
    cleanup_expired_import_files_task,
    generate_promo_codes_task,
    import_promo_codes_task,
)


def _xlsx_bytes(rows):
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


class PromoAdminTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            email='admin@example.com',
            password='test-password',
        )
        self.client.force_login(self.admin_user)

    def test_promo_admin_pages_render(self):
        urls = (
            reverse('admin:promo_promocode_changelist'),
            reverse('admin:promo_promocodeattempt_changelist'),
        )

        for url in urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)

    def test_generation_dialog_uses_relative_htmx_path(self):
        response = self.client.get(
            reverse('admin:promo_promocode_generate_codes'),
        )

        self.assertEqual(response.status_code, 200)
        body = response.content.decode()
        self.assertIn(
            'hx-post="/admin/promo/promocode/generate_codes/"',
            body,
        )
        self.assertNotIn('hx-post="http', body)

    @patch('promo.admin.generate_promo_codes_task.delay')
    def test_starts_promo_code_generation_from_admin(self, delay):
        delay.return_value = SimpleNamespace(id='celery-task-id')

        response = self.client.post(
            reverse('admin:promo_promocode_generate_codes'),
            {
                '_form_submitted': True,
                'count': 250_000,
            },
        )

        self.assertRedirects(
            response,
            reverse('admin:promo_promocodegeneration_changelist'),
        )
        generation = PromoCodeGeneration.objects.get()
        self.assertEqual(generation.requested_count, 250_000)
        self.assertEqual(generation.created_by, self.admin_user)
        self.assertEqual(generation.celery_task_id, 'celery-task-id')
        delay.assert_called_once_with(generation.pk)

    @patch('promo.admin.generate_promo_codes_task.delay')
    def test_generation_dialog_redirects_to_generation_list(self, delay):
        delay.return_value = SimpleNamespace(id='celery-task-id')

        response = self.client.post(
            reverse('admin:promo_promocode_generate_codes'),
            {
                '_form_submitted': True,
                'count': 250_000,
            },
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 204)
        self.assertEqual(
            response.headers['HX-Redirect'],
            reverse('admin:promo_promocodegeneration_changelist'),
        )
        self.assertEqual(PromoCodeGeneration.objects.count(), 1)

    @patch('promo.admin.generate_promo_codes_task.delay')
    def test_rejects_second_active_generation(self, delay):
        PromoCodeGeneration.objects.create(
            requested_count=100,
            created_by=self.admin_user,
        )

        response = self.client.post(
            reverse('admin:promo_promocode_generate_codes'),
            {
                '_form_submitted': True,
                'count': 200,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(PromoCodeGeneration.objects.count(), 1)
        delay.assert_not_called()

    @patch('promo.admin.import_promo_codes_task.delay')
    def test_queues_xlsx_import_from_admin(self, delay):
        delay.return_value = SimpleNamespace(id='import-task-id')
        upload = SimpleUploadedFile(
            'codes.xlsx',
            _xlsx_bytes([('AB12CD34', 'ignored')]),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

        response = self.client.post(
            reverse('admin:promo_promocode_import_codes'),
            {'_form_submitted': True, 'file': upload},
        )

        import_job = PromoCodeImport.objects.get()
        self.assertRedirects(
            response,
            reverse('admin:promo_promocodeimport_changelist'),
        )
        self.assertEqual(import_job.original_filename, 'codes.xlsx')
        self.assertEqual(import_job.celery_task_id, 'import-task-id')
        delay.assert_called_once_with(import_job.pk)

    @patch('promo.admin.import_promo_codes_task.delay')
    def test_import_dialog_redirects_to_import_list(self, delay):
        delay.return_value = SimpleNamespace(id='import-task-id')
        upload = SimpleUploadedFile(
            'codes.xlsx',
            _xlsx_bytes([('AB12CD34',)]),
            content_type=(
                'application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sheet'
            ),
        )

        response = self.client.post(
            reverse('admin:promo_promocode_import_codes'),
            {'_form_submitted': True, 'file': upload},
            HTTP_HX_REQUEST='true',
        )

        self.assertEqual(response.status_code, 204)
        self.assertEqual(
            response.headers['HX-Redirect'],
            reverse('admin:promo_promocodeimport_changelist'),
        )
        self.assertEqual(PromoCodeImport.objects.count(), 1)

    @override_settings(XLSX_MAX_UPLOAD_SIZE=5)
    @patch('promo.admin.import_promo_codes_task.delay')
    def test_rejects_xlsx_over_size_limit(self, delay):
        upload = SimpleUploadedFile('codes.xlsx', b'123456')

        response = self.client.post(
            reverse('admin:promo_promocode_import_codes'),
            {'_form_submitted': True, 'file': upload},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'Размер файла не должен превышать 20 МБ.',
        )
        self.assertFalse(PromoCodeImport.objects.exists())
        delay.assert_not_called()


class PromoCodeImportTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        settings_override = override_settings(
            MEDIA_ROOT=self.media_directory.name,
        )
        settings_override.enable()
        self.addCleanup(settings_override.disable)
        self.addCleanup(self.media_directory.cleanup)
        self.admin_user = User.objects.create_superuser(
            email='importer@example.com',
            password='test-password',
        )

    def create_import(self, rows):
        upload = SimpleUploadedFile('codes.xlsx', _xlsx_bytes(rows))
        return PromoCodeImport.objects.create(
            source_file=upload,
            original_filename='codes.xlsx',
            created_by=self.admin_user,
        )

    def test_imports_first_column_and_writes_skipped_rows(self):
        PromoCode.objects.create(code='EXIST001')
        import_job = self.create_import(
            (
                ('NEWCODE1', 'ignored'),
                ('newcode2', 'ignored'),
                ('NEWCODE1', 'ignored'),
                ('invalid', 'ignored'),
                ('EXIST001', 'ignored'),
                (None, 'ignored'),
            )
        )

        result = import_promo_codes_task(import_job.pk)

        import_job.refresh_from_db()
        self.assertEqual(import_job.status, PromoCodeImport.Status.COMPLETED)
        self.assertEqual(import_job.processed_count, 5)
        self.assertEqual(import_job.imported_count, 2)
        self.assertEqual(import_job.skipped_count, 3)
        self.assertEqual(result['imported_count'], 2)
        self.assertTrue(PromoCode.objects.filter(code='NEWCODE1').exists())
        self.assertTrue(PromoCode.objects.filter(code='NEWCODE2').exists())
        self.assertTrue(import_job.error_file)

        with import_job.error_file.open('rb') as error_file:
            workbook = load_workbook(error_file, read_only=True)
            rows = list(workbook.active.iter_rows(values_only=True))
            workbook.close()
        reasons = {row[2] for row in rows[1:]}
        self.assertEqual(
            reasons,
            {
                'Дубликат в файле',
                'Неверный формат',
                'Код уже существует в базе',
            },
        )

    @override_settings(GENERATED_FILE_RETENTION_DAYS=7)
    def test_cleanup_deletes_files_but_keeps_import_history(self):
        import_job = self.create_import((('NEWCODE1',),))
        import_promo_codes_task(import_job.pk)
        PromoCodeImport.objects.filter(pk=import_job.pk).update(
            created_at=timezone.now() - timedelta(days=8),
        )

        deleted_count = cleanup_expired_import_files_task()

        import_job.refresh_from_db()
        self.assertEqual(deleted_count, 1)
        self.assertFalse(import_job.source_file)
        self.assertTrue(PromoCodeImport.objects.filter(pk=import_job.pk).exists())


class PromoCodeGenerationTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            email='generator@example.com',
            password='test-password',
        )

    @patch('promo.services.generation.BATCH_SIZE', 3)
    @patch('promo.services.generation._generate_unique_candidates')
    def test_adds_exact_requested_count_despite_existing_codes(
        self,
        generate_candidates,
    ):
        PromoCode.objects.create(code='AAAAAAAA')
        generate_candidates.side_effect = (
            {'AAAAAAAA', 'BBBBBBBB', 'CCCCCCCC'},
            {'DDDDDDDD'},
        )
        generation = PromoCodeGeneration.objects.create(
            requested_count=3,
            created_by=self.admin_user,
        )

        result = generate_promo_codes_task(generation.pk)

        generation.refresh_from_db()
        self.assertEqual(result, 3)
        self.assertEqual(generation.generated_count, 3)
        self.assertEqual(
            generation.status,
            PromoCodeGeneration.Status.COMPLETED,
        )
        self.assertEqual(PromoCode.objects.count(), 4)

    @patch(
        'promo.tasks.generate_promo_codes',
        side_effect=RuntimeError('Database unavailable'),
    )
    def test_marks_generation_as_failed(self, generate_codes):
        generation = PromoCodeGeneration.objects.create(
            requested_count=3,
            created_by=self.admin_user,
        )

        with self.assertRaisesRegex(RuntimeError, 'Database unavailable'):
            generate_promo_codes_task(generation.pk)

        generation.refresh_from_db()
        self.assertEqual(generation.status, PromoCodeGeneration.Status.FAILED)
        self.assertEqual(generation.error, 'Database unavailable')
        self.assertIsNotNone(generation.finished_at)


class PromoCodeConcurrencyTests(TransactionTestCase):
    register_url = reverse('promo-code-register')

    def create_participant(self, email):
        user = User.objects.create_user(
            email=email,
            password='StrongPassword_123!',
            is_email_verified=True,
        )
        Profile.objects.create(
            user=user,
            first_name='Михаил',
            last_name='Иванов',
            no_middle_name=True,
            phone='+7 (999) 123-45-67',
            personal_data_consent_at=timezone.now(),
            promo_code_email_notifications=False,
        )
        return user, create_token_pair(user)['access']

    def register_concurrently(self, requests):
        barrier = Barrier(len(requests))

        def send_request(token, code):
            connections.close_all()
            try:
                barrier.wait(timeout=5)
                response = Client().post(
                    self.register_url,
                    {'code': code},
                    content_type='application/json',
                    HTTP_AUTHORIZATION=f'Bearer {token}',
                )
                return response.status_code, response.json()
            finally:
                connections.close_all()

        with ThreadPoolExecutor(max_workers=len(requests)) as executor:
            futures = [
                executor.submit(send_request, token, code)
                for token, code in requests
            ]
            return [future.result(timeout=10) for future in futures]

    def test_only_one_user_registers_code_under_concurrent_requests(self):
        first_user, first_token = self.create_participant(
            'first@example.com'
        )
        second_user, second_token = self.create_participant(
            'second@example.com'
        )
        promo_code = PromoCode.objects.create(code='AB12CD34')

        responses = self.register_concurrently(
            (
                (first_token, promo_code.code),
                (second_token, promo_code.code),
            )
        )

        self.assertEqual(sorted(status for status, _ in responses), [201, 400])
        failed_response = next(body for status, body in responses if status == 400)
        self.assertEqual(
            failed_response['reason'],
            PromoCodeAttempt.Reason.ALREADY_REGISTERED,
        )
        promo_code.refresh_from_db()
        self.assertIn(promo_code.registered_by_id, (first_user.pk, second_user.pk))
        self.assertEqual(
            PromoCodeAttempt.objects.filter(
                result=PromoCodeAttempt.Result.SUCCESS,
            ).count(),
            1,
        )
        self.assertEqual(PromoCodeAttempt.objects.count(), 2)

    def test_concurrent_failures_trigger_single_rate_limit_ban(self):
        _, token = self.create_participant('limited@example.com')

        responses = self.register_concurrently(
            tuple(
                (token, code)
                for code in ('AA11AA11', 'BB22BB22', 'CC33CC33', 'DD44DD44')
            )
        )

        self.assertEqual(
            sorted(status for status, _ in responses),
            [400, 400, 400, 429],
        )
        blocked_response = next(
            body for status, body in responses if status == 429
        )
        self.assertEqual(
            blocked_response['reason'],
            PromoCodeAttempt.Reason.RATE_LIMIT,
        )
        self.assertGreater(blocked_response['retry_after'], 0)
        self.assertLessEqual(blocked_response['retry_after'], 300)
        self.assertEqual(
            PromoCodeAttempt.objects.filter(
                result=PromoCodeAttempt.Result.FAILURE,
            ).count(),
            3,
        )
        self.assertEqual(
            PromoCodeAttempt.objects.filter(
                result=PromoCodeAttempt.Result.BLOCKED,
                reason=PromoCodeAttempt.Reason.RATE_LIMIT,
            ).count(),
            1,
        )

    def test_only_one_concurrent_generation_can_be_active(self):
        admin_user = User.objects.create_superuser(
            email='admin@example.com',
            password='test-password',
        )
        barrier = Barrier(2)

        def create_generation():
            connections.close_all()
            try:
                barrier.wait(timeout=5)
                PromoCodeGeneration.objects.create(
                    requested_count=100,
                    created_by_id=admin_user.pk,
                )
                return True
            except IntegrityError:
                return False
            finally:
                connections.close_all()

        with ThreadPoolExecutor(max_workers=2) as executor:
            results = [
                future.result(timeout=10)
                for future in (
                    executor.submit(create_generation),
                    executor.submit(create_generation),
                )
            ]

        self.assertEqual(sorted(results), [False, True])
        self.assertEqual(PromoCodeGeneration.objects.count(), 1)


class PromoCodeApiTests(TestCase):
    list_url = reverse('promo-code-list')
    register_url = reverse('promo-code-register')
    status_url = reverse('promo-code-registration-status')

    def setUp(self):
        self.user = User.objects.create_user(
            email='user@example.com',
            password='StrongPassword_123!',
            is_email_verified=True,
        )
        self.profile = Profile.objects.create(
            user=self.user,
            first_name='Михаил',
            last_name='Иванов',
            no_middle_name=True,
            phone='+7 (999) 123-45-67',
            personal_data_consent_at=timezone.now(),
        )
        self.tokens = create_token_pair(self.user)

    def register(self, code):
        return self.client.post(
            self.register_url,
            {'code': code},
            content_type='application/json',
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )

    def test_requires_authentication(self):
        response = self.client.get(self.list_url)

        self.assertEqual(response.status_code, 401)

    @patch('promo.services.notifications.schedule_registration_email')
    def test_registers_existing_code(self, schedule_email):
        promo_code = PromoCode.objects.create(code='AB12CD34')

        with self.captureOnCommitCallbacks(execute=True):
            response = self.register('ab12cd34')

        self.assertEqual(response.status_code, 201)
        promo_code.refresh_from_db()
        self.assertEqual(promo_code.registered_by, self.user)
        self.assertIsNotNone(promo_code.registered_at)
        attempt = PromoCodeAttempt.objects.get()
        self.assertEqual(attempt.result, PromoCodeAttempt.Result.SUCCESS)
        schedule_email.assert_called_once_with(promo_code.pk)

    def test_requires_completed_profile(self):
        self.profile.first_name = ''
        self.profile.save()
        PromoCode.objects.create(code='AB12CD34')

        response = self.register('AB12CD34')

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()['reason'], 'profile_incomplete')
        self.assertEqual(
            PromoCodeAttempt.objects.get().result,
            PromoCodeAttempt.Result.BLOCKED,
        )

    def test_rejects_invalid_format_and_missing_code(self):
        invalid_response = self.register('ABC')
        missing_response = self.register('ZZ99ZZ99')

        self.assertEqual(invalid_response.status_code, 400)
        self.assertEqual(
            invalid_response.json()['reason'],
            PromoCodeAttempt.Reason.INVALID_FORMAT,
        )
        self.assertEqual(missing_response.status_code, 400)
        self.assertEqual(
            missing_response.json()['reason'],
            PromoCodeAttempt.Reason.NOT_FOUND,
        )

    def test_counts_empty_and_long_codes_as_failed_attempts(self):
        empty_response = self.register('')
        long_response = self.register('A' * 100)

        self.assertEqual(empty_response.status_code, 400)
        self.assertEqual(long_response.status_code, 400)
        self.assertEqual(
            PromoCodeAttempt.objects.filter(
                result=PromoCodeAttempt.Result.FAILURE,
                reason=PromoCodeAttempt.Reason.INVALID_FORMAT,
            ).count(),
            2,
        )
        self.assertEqual(
            PromoCodeAttempt.objects.order_by('-created_at')
            .first()
            .raw_code,
            'A' * 64,
        )

    def test_rejects_already_registered_code(self):
        other_user = User.objects.create_user(
            email='other@example.com',
            password='StrongPassword_123!',
        )
        PromoCode.objects.create(
            code='AB12CD34',
            registered_by=other_user,
            registered_at=timezone.now(),
        )

        response = self.register('AB12CD34')

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()['reason'], 'already_registered')

    def test_blocks_for_five_minutes_after_three_recent_failures(self):
        PromoCode.objects.create(code='AB12CD34')
        for code in ('AA11AA11', 'BB22BB22', 'CC33CC33'):
            self.assertEqual(self.register(code).status_code, 400)

        blocked_response = self.register('AB12CD34')
        repeated_response = self.register('AB12CD34')

        self.assertEqual(blocked_response.status_code, 429)
        self.assertEqual(blocked_response.json()['retry_after'], 300)
        self.assertEqual(repeated_response.status_code, 429)
        self.assertEqual(
            PromoCodeAttempt.objects.filter(
                result=PromoCodeAttempt.Result.BLOCKED,
                reason=PromoCodeAttempt.Reason.RATE_LIMIT,
            ).count(),
            1,
        )
        self.assertFalse(
            PromoCode.objects.get(code='AB12CD34').registered_by_id
        )

        status_response = self.client.get(
            self.status_url,
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )
        self.assertEqual(status_response.status_code, 200)
        self.assertTrue(status_response.json()['is_blocked'])
        self.assertGreater(status_response.json()['retry_after'], 0)
        self.assertLessEqual(status_response.json()['retry_after'], 300)
        self.assertIsNotNone(status_response.json()['blocked_until'])

    def test_registration_status_is_open_without_active_ban(self):
        response = self.client.get(
            self.status_url,
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {
                'is_blocked': False,
                'retry_after': 0,
                'blocked_until': None,
            },
        )

    def test_lists_only_current_user_codes(self):
        other_user = User.objects.create_user(
            email='other@example.com',
            password='StrongPassword_123!',
        )
        PromoCode.objects.create(
            code='AB12CD34',
            registered_by=self.user,
            registered_at=timezone.now(),
        )
        PromoCode.objects.create(
            code='EF56GH78',
            registered_by=other_user,
            registered_at=timezone.now(),
        )

        response = self.client.get(
            self.list_url,
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['count'], 1)
        self.assertEqual(response.json()['results'][0]['code'], 'AB12CD34')

    def test_lists_participating_lost_and_winning_statuses(self):
        completed_at = timezone.now() - timedelta(days=1)
        period_started_at = completed_at - timedelta(days=1)
        period_ended_at = completed_at + timedelta(minutes=1)
        draw = Draw.objects.create(
            draw_date=timezone.localdate(completed_at),
            status=Draw.Status.COMPLETED,
            completed_at=timezone.now(),
            period_started_at=period_started_at,
            period_ended_at=period_ended_at,
        )
        lost_code = PromoCode.objects.create(
            code='LOST1234',
            registered_by=self.user,
            registered_at=completed_at,
        )
        winning_code = PromoCode.objects.create(
            code='WON12345',
            registered_by=self.user,
            registered_at=completed_at,
        )
        current_time = timezone.now()
        Draw.objects.create(
            draw_date=timezone.localdate(current_time),
            status=Draw.Status.COMPLETED,
            completed_at=current_time - timedelta(hours=1),
            period_started_at=period_ended_at,
            period_ended_at=current_time - timedelta(hours=1),
            trigger=Draw.Trigger.MANUAL,
        )
        participating_code = PromoCode.objects.create(
            code='PLAY1234',
            registered_by=self.user,
            registered_at=current_time,
        )
        Winner.objects.create(
            draw=draw,
            user=self.user,
            promo_code=winning_code,
            prize=Winner.Prize.AIRPODS,
        )

        response = self.client.get(
            self.list_url,
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )

        self.assertEqual(response.status_code, 200)
        codes = {
            item['code']: item for item in response.json()['results']
        }
        self.assertEqual(codes[lost_code.code]['status'], 'not_won')
        self.assertIsNone(codes[lost_code.code]['prize'])
        self.assertEqual(codes[winning_code.code]['status'], 'won')
        self.assertEqual(
            codes[winning_code.code]['prize'],
            {'code': Winner.Prize.AIRPODS, 'name': 'AirPods'},
        )
        self.assertEqual(
            codes[participating_code.code]['status'],
            'participating',
        )

    def test_paginates_user_codes_by_ten(self):
        PromoCode.objects.bulk_create(
            [
                PromoCode(
                    code=f'CD{index:06d}',
                    registered_by=self.user,
                    registered_at=timezone.now(),
                )
                for index in range(12)
            ]
        )

        first_page = self.client.get(
            self.list_url,
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )
        second_page = self.client.get(
            f'{self.list_url}?page=2',
            HTTP_AUTHORIZATION=f"Bearer {self.tokens['access']}",
        )

        self.assertEqual(first_page.status_code, 200)
        self.assertEqual(first_page.json()['count'], 12)
        self.assertEqual(len(first_page.json()['results']), 10)
        self.assertIsNotNone(first_page.json()['next'])
        self.assertEqual(second_page.status_code, 200)
        self.assertEqual(len(second_page.json()['results']), 2)
        self.assertIsNone(second_page.json()['next'])

    @override_settings(
        EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend',
    )
    def test_registration_email_respects_profile_setting(self):
        promo_code = PromoCode.objects.create(
            code='AB12CD34',
            registered_by=self.user,
            registered_at=timezone.now(),
        )

        send_registration_email(promo_code.pk)
        self.assertEqual(len(mail.outbox), 1)

        self.profile.promo_code_email_notifications = False
        self.profile.save()
        send_registration_email(promo_code.pk)
        self.assertEqual(len(mail.outbox), 1)
